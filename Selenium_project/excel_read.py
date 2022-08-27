import openpyxl

path = "D:\\Selenium_project\\EXCEL_DATA.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  # This loads the active sheet.


rows = sheet.max_row
cols = sheet.max_column

print(rows, cols)

# Now print the rows and cols.
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end=' ')
    print()


